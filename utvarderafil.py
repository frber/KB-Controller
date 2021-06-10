import os
import openpyxl




class UtvarderaFil:

    def __init__(self, root, fil):
        self.fil = fil
        self.filplats = os.path.join(root, fil).replace("\\","/")

    def returnera_filplats(self):
        return self.filplats

    def avgor_om_excel(self):
       if self.filplats.endswith(".xlsx") or self.filplats.endswith(".xlsm"):
           return True

    def storlek(self):
        try:
            storlek = os.path.getsize(self.filplats)
            if storlek < 700000:
                return True
        except:
            return False

    def avgor_om_berper(self):
        if self.avgor_om_excel() and self.storlek():
            try:
                wb = openpyxl.load_workbook(self.filplats, data_only=True)
            except:
                return "fel"
            for sheet in wb.worksheets:
                cell_kst = sheet.cell(31, 7).value
                cell_projledare = sheet.cell(34, 7).value
                if isinstance(cell_kst, str) == True and isinstance(cell_projledare, str) == True:
                    if cell_kst.lower() == "kst" and cell_projledare.lower() == "projektledare":
                        wb.close()
                        return sheet
            wb.close()

    def filnamn_clean(self):
        filnamn_clean = os.path.splitext(self.fil)[0]
        return filnamn_clean