import openpyxl
from outputmodel import OutputModel
import datetime

class KontrolleraBerper:

    def __init__(self, sheet, filnamn_clean, filplats):
        self.sheet = sheet
        self.filnamn_clean = filnamn_clean
        self.filplats = filplats

    def validera(self):
        data = OutputModel()
        data.filnamn_clean = self.filnamn_clean
        data.filplats = self.filplats
        check = 0
        for row in self.sheet['F30:F100']:
            for cell in row:
                if self.kontroll_periodisering(cell):
                    data.kontroll_per = 1
                if self.kontroll_anlaggning(cell):
                    data.kontroll_anl = 1
                if self.hamta_resultat(cell) != None:
                    data.resultat = self.hamta_resultat(cell)
                if self.hamta_periodsiering(cell) != None:
                    data.periodisering = self.hamta_periodsiering(cell)
                if self.kontrollera_kommentar(cell) != None:
                    data.kommentar = self.kontrollera_kommentar(cell)
                if self.kontrollera_perkonto(cell):
                    data.konto_per = 1
                if self.kontrollera_belopp(cell):
                    data.belopp = 1
                if self.hamta_kst() != None:
                    data.kst = self.hamta_kst()
                if self.hamta_projnummer() != None:
                    data.projekt = self.hamta_projnummer()
                if self.hamta_perkonton(cell):
                    check += 1
                    anv_perkonto_belopp = self.hamta_perkonton(cell)
                if self.hamta_intakt(cell) != None:
                    data.intakt = self.hamta_intakt(cell)
                if self.hamta_kostnad(cell) != None:
                    data.kostnad = self.hamta_kostnad(cell)





        if check > 0:
            data.anv_perkonto_belopp = self.hantera_perkonton(anv_perkonto_belopp)
        else:
            data.anv_perkonto_belopp = 0

        konto_per = data.konto_per
        belopp = data.belopp
        periodisering = data.periodisering
        intakt = data.kostnad
        kostnad = data.kostnad

        if self.kontrollera_laddning(konto_per, belopp, periodisering):
            data.kontroll_ladd = 1

        if self.int_kost_per_noll(intakt, kostnad, periodisering):
            data.int_kost_per_noll = 1
            #print(data.int_kost_per_noll)
        if self.kontrollera_slutdatum():
            data.slutdatum = 1
            #print(data.slutdatum)


        data.upprattad_av = self.hamta_upprattad_av()

        #print(self.hamta_slutdatum())
        return data

    def kontroll_periodisering(self, cell):
        # Kollar kontrollruta för periodisering i berper. Denna cell baseras på en excel-formel i filen, om den visar fel = True
        if cell.value == "Summa":
            if cell.offset(column=3).value == "Kontrollera din periodisering":
                return True

    def kontroll_anlaggning(self, cell):
        # Denna cell baseras på en excel-formel i filen. Denna får inte vara mer eller mindre än 0. Om det är så = True
        if cell.value == "Summa":
            anl = cell.offset(column=8).value
            if anl == None or isinstance(anl, str) == True:
                pass
            else:
                if anl > 0.01 or anl < -0.01:
                    return True

    def hamta_resultat(self, cell):
        if cell.value == "Resultat":
            return cell.offset(column=7).value

    def hamta_periodsiering(self, cell):
        if cell.value == "Årets periodisering":
            periodisering = cell.offset(column=7).value
            if isinstance(periodisering, int) == True or isinstance(periodisering, float) == True:
                # Vissa periodiseringar är - och vissa + eftersom ekonomer är som dom är. Vänder alla till + för logisk filtreringsmöjlighet.
                if periodisering < 0:
                    periodisering = -periodisering
                    return periodisering
                else:
                    return periodisering

    def kontrollera_kommentar(self, cell):
        # Det ska alltid finnans en kommentar i berper. Om det inte finns någon kommentar = 1. Annars returnera kommentar.
        if cell.value == "Förklaring till projektets genomförande/resultat. Ekonomiskt och tidsmässigt mm. Ange om bidrag inbetalt till annat projekt.":
            lista_kommentar = []
            kolumn = 0
            kolumn2 = 0
            while kolumn < 11:
                kommentar = cell.offset(row=1, column=kolumn).value
                kolumn += 1
                if kommentar != None:
                    lista_kommentar.append(str(kommentar))
            while kolumn2 < 11:
                kommentar2 = cell.offset(row=2, column=kolumn2).value
                kolumn2 += 1
                if kommentar2 != None:
                    lista_kommentar.append(str(kommentar2))
            if not lista_kommentar:
                return 1
            else:
                lista_kommentar = ' '.join(lista_kommentar)
                return lista_kommentar

    def kontrollera_perkonto(self, cell):
        # Om ett periodiseringskonto har använts, returnera True. Detta används sedan som ett argument i kontrollera_laddning().
        if cell.value == "Konto":
            rad = 0
            for x in range(16):
                rad += 1
                # print(rad)
                konto = cell.offset(row=rad).value
                if konto == None or isinstance(konto, str) == True:
                    pass
                else:
                    if konto >= 1000 and konto <= 1632 or konto >= 1635 and konto <= 2732 or konto >= 2732 and konto < 3000:
                        return True

    def kontrollera_belopp(self, cell):
        # Om det finns ett belopp, returnera True. Detta används sedan som ett argument i kontrollera_laddning().
        if cell.value == "Konto":
            rad = 0
            for x in range(16):
                rad += 1
                belopp = cell.offset(row=rad, column=9).value
                if belopp == None or isinstance(belopp, str) == True:
                    pass
                else:
                    if belopp > 0 or belopp < 0:
                        return True

    def kontrollera_laddning(self, konto_per, belopp, periodisering):
        # Om man har skrivit in ett periodiseringskonto och ett belopp, men berper inte visar någon inladdad periodisering för nuvarande år, är sannolikheten hög att man har glömt att ladda berper. Om så är fallet = True
        if konto_per == 1 and belopp == 1 and periodisering == 0:
            return True

    def hamta_kst(self):
        kst = str(self.sheet.cell(32, 8).value)[:3]
        return kst

    def hamta_projnummer(self):
        projekt = self.sheet.cell(32, 8).value
        return projekt

    def hamta_perkonton(self, cell):
        # Hämtar periodiseringskonton och belopp (de som är inladdade från Agresso i berper). Listan som kommer av detta används sedan i hantera_perkonton().
        lista_perkonto_belopp = []
        if cell.value == "Specifikation av årets periodisering, KR":
            rad = 0
            for x in range(10):
                rad += 1
                perkonto = cell.offset(row=rad, column=3).value
                if perkonto != None and isinstance(perkonto, (int, float)):
                    belopp = cell.offset(row=rad, column=2).value
                    lista_perkonto_belopp.append([perkonto, belopp])

        return lista_perkonto_belopp

    def hantera_perkonton(self, anv_perkonto_belopp):
        #Summerar alla belopp där samma konton har används och returnerar en "lista av lista" på detta.
        lista_konton = []
        for x in anv_perkonto_belopp:
            lista_konton.append(x[0])


        konton = set(lista_konton)
        lista_sum_konton = []

        for y in konton:
            tot = 0
            for z in anv_perkonto_belopp:
                if y == z[0]:
                    if isinstance(z[1], str) == False:
                        if z[1] != None:
                            tot += z[1]
            lista_sum_konton.append([y, tot])
        return lista_sum_konton

    def hamta_upprattad_av(self):
        return self.sheet.cell(36, 8).value

    def hamta_intakt(self, cell):
        if cell.value != None and isinstance(cell.value, str):
            if cell.value.lower() == "summa intäkter":
                intakt = cell.offset(column=7).value
                return intakt

    def hamta_kostnad(self, cell):
        if cell.value != None and isinstance(cell.value, str):
            if cell.value.lower() == "summa kostnader":
                kostnad = cell.offset(column=7).value
                return kostnad


    def int_kost_per_noll(self, intakt, kostnad, periodisering):
        if intakt+kostnad == 0 and periodisering > 0:
            return True

    def kontrollera_slutdatum(self):
        idag = datetime.datetime.now()
        slutdatum =  self.sheet.cell(33, 11).value


        if isinstance(slutdatum, datetime.datetime):
            if slutdatum < idag:
                return True

        # Om en berper buggar och det står 1900-01-01, då blir formatet tid, detta fångar upp det.
        if isinstance(slutdatum, datetime.time):
            return True


