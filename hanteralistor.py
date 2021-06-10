import openpyxl


class HanteraListor:

    def __init__(self, lista_alla_filer, lista_berpers, lista_trasiga_filer, lista_for_stora_excelfiler, wb_berperdata):
        self.lista_alla_filer = lista_alla_filer
        self.lista_berpers = lista_berpers
        self.lista_trasiga_filer = lista_trasiga_filer
        self.list_for_stora_excelfiler = lista_for_stora_excelfiler
        self.wb_berperdata = wb_berperdata
        self.ws_ejberper = self.wb_berperdata["Filer som ej bedöms vara berper"]
        self.ws_trasiga = self.wb_berperdata['Excelfiler som ej kunde öppnas']
        self.ws_forstora = self.wb_berperdata['För stora excelfiler']


    def kontrollera_listor(self):
        # Filer som ej bedöms vara berpers med sortering av filformat.
        diff_alla_filer = set(self.lista_alla_filer) - set(self.lista_berpers)
        diff_alla_filer = list(diff_alla_filer)
        diff_alla_filer.sort(reverse=True, key=lambda x: x[-4:])

        # Förhindrar att en fil som är trasig dyker upp i listan för filer som ej bedöms vara berper
        for x in diff_alla_filer:
            if x not in self.lista_trasiga_filer:
                self.ws_ejberper.cell(row=self.ws_ejberper.max_row + 1, column=1).value = x

        for y in self.lista_trasiga_filer:
            self.ws_trasiga.cell(row=self.ws_trasiga.max_row  + 1, column=1).value = y

        for z in self.list_for_stora_excelfiler:
            self.ws_forstora.cell(row=self.ws_forstora.max_row + 1, column=1).value = z

        self.wb_berperdata.close()