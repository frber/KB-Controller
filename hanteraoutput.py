import os
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Protection
import pandas as pd



class HanteraOutput:

    def __init__(self, berper, wb_berperdata, berperdata):
        self.berper = berper
        self.wb_berperdata = wb_berperdata
        self.berperdata = berperdata
        self.ws_avvikelser = self.wb_berperdata["Avvikelser"]
        #self.ws_avvikelser2 = self.wb_berperdata["Avvikelser2"]
        self.ws_kommentar = self.wb_berperdata["Alla Berper"]
        self.ws_ejberper = self.wb_berperdata["Filer som ej bedöms vara berper"]
        self.ws_trasiga = self.wb_berperdata['Excelfiler som ej kunde öppnas']
        self.ws_forstora = self.wb_berperdata['För stora excelfiler']


    def skriv_output(self):

        self.filnamn_clean = self.berper.filnamn_clean
        self.filplats = self.berper.filplats
        self.kontroll_per = self.berper.kontroll_per
        self.kontroll_anl = self.berper.kontroll_anl
        self.resultat = self.berper.resultat
        self.periodisering = self.berper.periodisering
        self.kommentar = self.berper.kommentar
        self.laddning = self.berper.kontroll_ladd
        self.belopp = self.berper.belopp
        self.kst = self.berper.kst
        self.projekt = self.berper.projekt
        self.anv_perkonto_belopp = self.berper.anv_perkonto_belopp
        self.upprattad_av = self.berper.upprattad_av
        self.int_kost_per_noll = self.berper.int_kost_per_noll
        self.slutdatum = self.berper.slutdatum


        self.alla_berper()
        self.avvikelser()
        #self.avvikelser_tva()




    def alla_berper(self):

        #KST
        out_kst = self.ws_kommentar.cell(row=self.ws_kommentar.max_row+1, column=1)
        out_kst.value = self.kst
        out_kst.alignment = Alignment(horizontal='center')
        #PROJEKT
        out_projekt = self.ws_kommentar.cell(row=self.ws_kommentar.max_row, column=2)
        out_projekt.value = self.projekt
        out_projekt.alignment = Alignment(horizontal='center')
        #FILNAMN
        out_filnamn = self.ws_kommentar.cell(row=self.ws_kommentar.max_row, column=3)
        out_filnamn.value = self.filnamn_clean
        out_filnamn.hyperlink = self.filplats
        out_filnamn.style = "Hyperlink"
        out_filnamn.alignment = Alignment(horizontal='center')
        #TOTAL PERIODISERING
        out_per = self.ws_kommentar.cell(row=self.ws_kommentar.max_row, column=4)
        out_per.value = self.periodisering
        out_per.alignment = Alignment(horizontal='center')
        out_per.number_format = '#,##0.00'
        #UPPRÄTTAD AV
        out_upp = self.ws_kommentar.cell(row=self.ws_kommentar.max_row, column=5)
        out_upp.value = self.upprattad_av
        out_upp.alignment = Alignment(horizontal='center')
        #KOMMENTAR
        out_kom = self.ws_kommentar.cell(row=self.ws_kommentar.max_row, column=6)
        if self.kommentar == 1:
            out_kom.value = "KOMMENTAR SAKNAS"
        else:
            out_kom.value = self.kommentar
        #LÄNK
        out_lank = self.ws_kommentar.cell(row=self.ws_kommentar.max_row, column=7)
        out_lank.value = self.filplats

    def avvikelser(self):

        if self.kontroll_per+self.kontroll_anl+self.laddning+self.int_kost_per_noll+self.slutdatum > 0 or self.kommentar == 1:
            #KST
            out_kst = self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row+1, column=1)
            out_kst.value = self.kst
            out_kst.alignment = Alignment(horizontal='center')
            #PROJEKT
            out_projekt =  self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row, column=2)
            out_projekt.value = self.projekt
            out_projekt.alignment = Alignment(horizontal='center')
            #FILNAMN
            out_filnamn = self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row, column=3)
            out_filnamn.value = self.filnamn_clean
            out_filnamn.hyperlink = self.filplats
            out_filnamn.style = "Hyperlink"
            out_filnamn.alignment = Alignment(horizontal='center')
            #PERIODISERING
            out_per = self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row, column=4)
            out_per.value = self.periodisering
            out_per.alignment = Alignment(horizontal='center')
            out_per.number_format = '#,##0.00'
            #BERPER UPPRÄTTAD AV
            out_upp = self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row, column=11)
            out_upp.value = self.upprattad_av
            out_upp.alignment = Alignment(horizontal='center')

            if self.kontroll_per == 1:
                out_kontroll_per = self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row, column=5)
                out_kontroll_per.value = "Kontrollera"
                out_kontroll_per.alignment = Alignment(horizontal='center')
            if self.kontroll_anl == 1:
                out_kontroll_anl = self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row, column=6)
                out_kontroll_anl.value = "Kontrollera"
                out_kontroll_anl.alignment = Alignment(horizontal='center')
            if self.laddning == 1:
                out_kontroll_ladd = self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row, column=7)
                out_kontroll_ladd.value = "Kontrollera"
                out_kontroll_ladd.alignment = Alignment(horizontal='center')
            if self.int_kost_per_noll == 1:
                out_int_kost = self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row, column=8)
                out_int_kost.value = "Kontrollera"
                out_int_kost.alignment = Alignment(horizontal='center')
            if self.slutdatum == 1:
                out_slutdat = self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row, column=9)
                out_slutdat.value = "Kontrollera"
                out_slutdat.alignment = Alignment(horizontal='center')
            if self.kommentar == 1:
                out_kontroll_kom = self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row, column=10)
                out_kontroll_kom.value = "Kontrollera"
                out_kontroll_kom.alignment = Alignment(horizontal='center')
            if self.kommentar != 1:
                out_kontroll_kom2 = self.ws_avvikelser.cell(row=self.ws_avvikelser.max_row, column=12)
                out_kontroll_kom2.value = self.kommentar



    def avvikelser_tva(self):

        if self.anv_perkonto_belopp != 0:
            for x in self.anv_perkonto_belopp:
                if self.kontroll_per+self.kontroll_anl+self.laddning+self.int_kost_per_noll+self.slutdatum > 0 or self.kommentar == 1:
                    #KST
                    out_kst = self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row+1, column=1)
                    out_kst.value = self.kst
                    out_kst.alignment = Alignment(horizontal='center')
                    #PROJEKT
                    out_projekt =  self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=2)
                    out_projekt.value = self.projekt
                    out_projekt.hyperlink = self.filplats
                    out_projekt.style = "Hyperlink"
                    out_projekt.alignment = Alignment(horizontal='center')
                    #PERIODISERINGSKONTO
                    period_konto = self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=3)
                    period_konto.value = x[0]
                    period_konto.alignment = Alignment(horizontal='center')
                    #BELOPP PERIODISERING
                    out_per = self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=4)
                    out_per.value = x[1]
                    out_per.alignment = Alignment(horizontal='center')
                    out_per.number_format = '#,##0.00'
                    #BERPER UPPRÄTTAD AV
                    out_upp = self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=5)
                    out_upp.value = self.upprattad_av
                    out_upp.alignment = Alignment(horizontal='center')
                    #LÄNK
                    out_lank = self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=13)
                    out_lank.value = self.filplats

                    if self.kontroll_per == 1:
                        out_kontroll_per = self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=6)
                        out_kontroll_per.value = "Kontrollera"
                        out_kontroll_per.alignment = Alignment(horizontal='center')
                    if self.kontroll_anl == 1:
                        out_kontroll_anl =  self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=7)
                        out_kontroll_anl.value = "Kontrollera"
                        out_kontroll_anl.alignment = Alignment(horizontal='center')
                    if self.laddning == 1:
                        out_kontroll_ladd = self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=8)
                        out_kontroll_ladd.value = "Kontrollera"
                        out_kontroll_ladd.alignment = Alignment(horizontal='center')

                    if self.int_kost_per_noll == 1:
                        out_int_kost =  self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=9)
                        out_int_kost.value = "Kontrollera"
                        out_int_kost.alignment = Alignment(horizontal='center')
                    if self.slutdatum == 1:
                        out_slutdat = self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=10)
                        out_slutdat.value = "Kontrollera"
                        out_slutdat.alignment = Alignment(horizontal='center')

                    if self.kommentar == 1:
                        out_kontroll_kom = self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=11)
                        out_kontroll_kom.value = "Kontrollera"
                        out_kontroll_kom.alignment = Alignment(horizontal='center')

                    if self.kommentar != 1:
                        out_kontroll_kom2 = self.ws_avvikelser2.cell(row=self.ws_avvikelser2.max_row, column=12)
                        out_kontroll_kom2.value = self.kommentar


    def avvikade_berper(self):
        if self.kontroll_per+self.kontroll_anl+self.laddning > 0 or self.kommentar == 1:
            return True

    def starta(self):
        os.startfile(self.berperdata)

